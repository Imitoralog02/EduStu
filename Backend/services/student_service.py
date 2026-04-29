from typing import Optional
from sqlalchemy.orm import Session
from sqlalchemy import or_
from fastapi import HTTPException

from models.student import Student, StudentStatusLog
from models.grade import Grade
from models.course import Course
from models.document import StudentDocument
from models.tuition import Tuition
from utils.grade_calc import calc_tong_ket


def _batch_gpa(db: Session, mssv_list: list[str]) -> dict[str, Optional[float]]:
    if not mssv_list:
        return {}
    rows = (
        db.query(Grade, Course.so_tin_chi)
        .join(Course, Grade.ma_hp == Course.ma_hp)
        .filter(Grade.mssv.in_(mssv_list))
        .all()
    )
    data: dict[str, tuple[float, int]] = {}
    for grade, tc in rows:
        tong = calc_tong_ket(grade.diem_gk, grade.diem_ck)
        if tong is not None:
            prev = data.get(grade.mssv, (0.0, 0))
            data[grade.mssv] = (prev[0] + tong * tc, prev[1] + tc)

    return {
        mssv: (round(data[mssv][0] / data[mssv][1], 2) if data.get(mssv, (0, 0))[1] > 0 else None)
        for mssv in mssv_list
    }


def _compute_gpa(db: Session, mssv: str) -> Optional[float]:
    return _batch_gpa(db, [mssv]).get(mssv)


def _to_out(student: Student, gpa: Optional[float] = None) -> dict:
    data = {c.name: getattr(student, c.name) for c in student.__table__.columns}
    data["ngay_sinh"] = str(data["ngay_sinh"]) if data.get("ngay_sinh") else None
    data["gpa"] = gpa
    return data


def list_students(
    db: Session,
    search: Optional[str],
    khoa: Optional[str],
    trang_thai: Optional[str],
    lop: Optional[str],
    page: int,
    page_size: int,
    nam_nhap_hoc: Optional[int] = None,
    thieu_giay_to: bool = False,
    no_hoc_phi: bool = False,
) -> dict:
    q = db.query(Student)
    if search:
        q = q.filter(
            or_(Student.mssv.ilike(f"%{search}%"), Student.ho_ten.ilike(f"%{search}%"))
        )
    if khoa:
        q = q.filter(Student.khoa == khoa)
    if trang_thai:
        q = q.filter(Student.trang_thai == trang_thai)
    if lop:
        q = q.filter(Student.lop.ilike(f"%{lop}%"))
    if nam_nhap_hoc:
        q = q.filter(Student.nam_nhap_hoc == nam_nhap_hoc)
    if thieu_giay_to:
        subq = (
            db.query(StudentDocument.mssv)
            .filter(StudentDocument.da_nop == False)
            .distinct()
            .subquery()
        )
        q = q.filter(Student.mssv.in_(subq))
    if no_hoc_phi:
        subq = (
            db.query(Tuition.mssv)
            .filter(Tuition.da_nop < Tuition.phai_nop)
            .subquery()
        )
        q = q.filter(Student.mssv.in_(subq))

    total = q.count()
    items = q.offset((page - 1) * page_size).limit(page_size).all()
    gpa_map = _batch_gpa(db, [sv.mssv for sv in items])
    return {"items": [_to_out(sv, gpa_map.get(sv.mssv)) for sv in items], "total": total}


def get_student(db: Session, mssv: str) -> dict:
    sv = db.query(Student).filter(Student.mssv == mssv).first()
    if not sv:
        raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên")
    return _to_out(sv, _compute_gpa(db, mssv))


def create_student(db: Session, data: dict) -> dict:
    if db.query(Student).filter(Student.mssv == data["mssv"]).first():
        raise HTTPException(status_code=400, detail="MSSV đã tồn tại")
    sv = Student(**data)
    db.add(sv)
    db.commit()
    db.refresh(sv)
    return _to_out(sv, None)


def update_student(db: Session, mssv: str, data: dict, nguoi_thay_doi: str = None) -> dict:
    sv = db.query(Student).filter(Student.mssv == mssv).first()
    if not sv:
        raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên")
    old_tt = sv.trang_thai
    for field, value in data.items():
        setattr(sv, field, value)
    if "trang_thai" in data and data["trang_thai"] != old_tt:
        db.add(StudentStatusLog(
            mssv=mssv,
            trang_thai_cu=old_tt,
            trang_thai_moi=data["trang_thai"],
            ly_do=data.get("ly_do_doi_tt"),
            nguoi_thay_doi=nguoi_thay_doi,
        ))
    db.commit()
    db.refresh(sv)
    return _to_out(sv, _compute_gpa(db, mssv))


def delete_student(db: Session, mssv: str, nguoi_thay_doi: str = None) -> dict:
    sv = db.query(Student).filter(Student.mssv == mssv).first()
    if not sv:
        raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên")
    old_tt = sv.trang_thai
    sv.trang_thai = "Thôi học"
    if old_tt != "Thôi học":
        db.add(StudentStatusLog(
            mssv=mssv,
            trang_thai_cu=old_tt,
            trang_thai_moi="Thôi học",
            nguoi_thay_doi=nguoi_thay_doi,
        ))
    db.commit()
    return {"message": f"Đã cập nhật trạng thái sinh viên {mssv} thành 'Thôi học'"}


def get_status_history(db: Session, mssv: str) -> list:
    sv = db.query(Student).filter(Student.mssv == mssv).first()
    if not sv:
        raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên")
    logs = (
        db.query(StudentStatusLog)
        .filter(StudentStatusLog.mssv == mssv)
        .order_by(StudentStatusLog.thoi_gian.desc())
        .all()
    )
    return [
        {
            "id": log.id,
            "trang_thai_cu": log.trang_thai_cu,
            "trang_thai_moi": log.trang_thai_moi,
            "ly_do": log.ly_do,
            "nguoi_thay_doi": log.nguoi_thay_doi,
            "thoi_gian": str(log.thoi_gian),
        }
        for log in logs
    ]


def export_student_profile(db: Session, mssv: str) -> bytes:
    """Xuất toàn bộ hồ sơ 1 sinh viên ra file Excel nhiều sheet."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import date

    sv = db.query(Student).filter(Student.mssv == mssv).first()
    if not sv:
        raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên")

    wb = Workbook()

    # ── Sheet 1: Thông tin cá nhân ────────────────────────────────────
    ws = wb.active
    ws.title = "Thông tin"
    header_font  = Font(bold=True, size=11, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2563EB")
    label_font   = Font(bold=True, size=10)
    center       = Alignment(horizontal="center", vertical="center")
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = f"HỒ SƠ SINH VIÊN — {sv.ho_ten} ({sv.mssv})"
    title_cell.font  = Font(bold=True, size=13, color="0F172A")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    fields = [
        ("Mã sinh viên",       sv.mssv),
        ("Họ và tên",          sv.ho_ten),
        ("Ngày sinh",          str(sv.ngay_sinh) if sv.ngay_sinh else ""),
        ("Giới tính",          sv.gioi_tinh or ""),
        ("Khoa",               sv.khoa or ""),
        ("Lớp",                sv.lop or ""),
        ("Năm nhập học",       str(sv.nam_nhap_hoc) if sv.nam_nhap_hoc else ""),
        ("Trạng thái",         sv.trang_thai or ""),
        ("Email",              sv.email or ""),
        ("Số điện thoại",      sv.so_dien_thoai or ""),
        ("Địa chỉ",            sv.dia_chi or ""),
        ("Đối tượng ưu tiên",  sv.doi_tuong or ""),
        ("Họ tên cha",         sv.ho_ten_cha or ""),
        ("Họ tên mẹ",          sv.ho_ten_me or ""),
        ("SĐT phụ huynh",      sv.sdt_phu_huynh or ""),
    ]
    for i, (label, value) in enumerate(fields, start=2):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = label_font
        ws[f"B{i}"] = value
        for col in ["A", "B"]:
            ws[f"{col}{i}"].border = border
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40

    # ── Sheet 2: Bảng điểm ───────────────────────────────────────────
    ws2 = wb.create_sheet("Bảng điểm")
    headers2 = ["Học kỳ", "Mã HP", "Tên học phần", "Tín chỉ", "Điểm GK", "Điểm CK", "Tổng kết", "Kết quả"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    grade_rows = (
        db.query(Grade, Course)
        .join(Course, Grade.ma_hp == Course.ma_hp)
        .filter(Grade.mssv == mssv)
        .order_by(Grade.hoc_ky)
        .all()
    )
    for row_idx, (g, c) in enumerate(grade_rows, 2):
        tong = calc_tong_ket(g.diem_gk, g.diem_ck)
        ket_qua = ("Đạt" if tong >= 5.0 else "Rớt") if tong is not None else ""
        for col, val in enumerate([
            g.hoc_ky, g.ma_hp, c.ten_hp, c.so_tin_chi,
            g.diem_gk, g.diem_ck,
            round(tong, 2) if tong is not None else "", ket_qua,
        ], 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border = border
    for w, col in zip([14, 10, 28, 8, 10, 10, 10, 8], range(1, 9)):
        ws2.column_dimensions[ws2.cell(1, col).column_letter].width = w

    # ── Sheet 3: Học phí ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Học phí")
    tuition = db.query(Tuition).filter(Tuition.mssv == mssv).first()
    headers3 = ["Phải nộp", "Miễn giảm", "Lý do miễn giảm", "Đã nộp", "Còn thiếu", "Hạn nộp", "Trạng thái"]
    for col, h in enumerate(headers3, 1):
        c3 = ws3.cell(row=1, column=col, value=h)
        c3.font = header_font; c3.fill = header_fill; c3.alignment = center; c3.border = border
    if tuition:
        actual_due = max(0, (tuition.phai_nop or 0) - (tuition.mien_giam or 0))
        con_thieu  = max(0, actual_due - (tuition.da_nop or 0))
        if (tuition.da_nop or 0) >= actual_due:
            tt = "Đã nộp"
        elif tuition.han_nop and tuition.han_nop < date.today():
            tt = "Quá hạn"
        elif (tuition.da_nop or 0) > 0:
            tt = "Nộp thiếu"
        else:
            tt = "Chưa nộp"
        for col, val in enumerate([
            tuition.phai_nop, tuition.mien_giam, tuition.ly_do_mien_giam or "",
            tuition.da_nop, con_thieu, str(tuition.han_nop) if tuition.han_nop else "", tt,
        ], 1):
            ws3.cell(row=2, column=col, value=val).border = border
    for w in [14, 14, 30, 14, 14, 14, 12]:
        ws3.column_dimensions[ws3.cell(1, headers3.index(headers3[w - 14]) + 1 if w >= 14 else w).column_letter].width = w

    # ── Sheet 4: Giấy tờ ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Giấy tờ")
    headers4 = ["Loại giấy tờ", "Trạng thái", "Ngày nộp", "File đính kèm", "Ghi chú"]
    for col, h in enumerate(headers4, 1):
        c4 = ws4.cell(row=1, column=col, value=h)
        c4.font = header_font; c4.fill = header_fill; c4.alignment = center; c4.border = border
    docs = db.query(StudentDocument).filter(StudentDocument.mssv == mssv).all()
    for row_idx, d in enumerate(docs, 2):
        for col, val in enumerate([
            d.loai_giay,
            "Đã nộp" if d.da_nop else "Chưa nộp",
            str(d.ngay_nop) if d.ngay_nop else "",
            d.file_name or "",
            d.ghi_chu or "",
        ], 1):
            ws4.cell(row=row_idx, column=col, value=val).border = border
    for w, col in zip([30, 12, 12, 30, 30], range(1, 6)):
        ws4.column_dimensions[ws4.cell(1, col).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
